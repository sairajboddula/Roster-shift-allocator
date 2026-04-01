"""Export service - renders schedules to CSV and Excel."""
import csv
import io
from datetime import date

from sqlalchemy.orm import Session

from app.models.schedule import Schedule
from app.models.employee import Employee
from app.models.department import Department
from app.models.shift import Shift
from app.utils.logger import get_logger

logger = get_logger(__name__)


class ExportService:

    def __init__(self, db: Session, user_id: int):
        self.db = db
        self.user_id = user_id

    def export_csv(
        self,
        roster_type: str,
        start_date: date,
        end_date: date,
        batch_id: str | None = None,
    ) -> str:
        """Return UTF-8 CSV string of schedules for the given filters."""
        rows = self._query_schedules(roster_type, start_date, end_date, batch_id)
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([
            "Date", "Employee", "Role", "Department",
            "Shift", "Start", "End", "AI Score", "Reason", "Manual Override",
        ])
        for row in rows:
            writer.writerow([
                row["date"], row["employee_name"], row["role"],
                row["department_name"], row["shift_name"],
                row["start_time"], row["end_time"],
                f"{row['ai_score']:.3f}", row["reason"],
                "Yes" if row["is_manual_override"] else "No",
            ])
        return output.getvalue()

    def export_excel(
        self,
        roster_type: str,
        start_date: date,
        end_date: date,
        batch_id: str | None = None,
    ) -> bytes:
        """Return Excel workbook bytes."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl not installed. Run: pip install openpyxl")

        rows = self._query_schedules(roster_type, start_date, end_date, batch_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{roster_type.upper()} Roster"

        # Header
        headers = ["Date", "Employee", "Role", "Department", "Shift",
                   "Start", "End", "AI Score", "Reason", "Manual Override"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(fill_type="solid", fgColor="1E3A5F")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for row_idx, row in enumerate(rows, 2):
            ws.cell(row=row_idx, column=1, value=row["date"])
            ws.cell(row=row_idx, column=2, value=row["employee_name"])
            ws.cell(row=row_idx, column=3, value=row["role"])
            ws.cell(row=row_idx, column=4, value=row["department_name"])
            ws.cell(row=row_idx, column=5, value=row["shift_name"])
            ws.cell(row=row_idx, column=6, value=row["start_time"])
            ws.cell(row=row_idx, column=7, value=row["end_time"])
            ws.cell(row=row_idx, column=8, value=round(row["ai_score"], 3))
            ws.cell(row=row_idx, column=9, value=row["reason"])
            ws.cell(row=row_idx, column=10, value="Yes" if row["is_manual_override"] else "No")

        # Auto-width
        for column_cells in ws.columns:
            max_len = max(len(str(c.value or "")) for c in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max_len + 4, 60)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _query_schedules(
        self,
        roster_type: str,
        start_date: date,
        end_date: date,
        batch_id: str | None,
    ) -> list[dict]:
        from sqlalchemy import and_

        q = (
            self.db.query(Schedule, Employee, Department, Shift)
            .join(Employee, Schedule.employee_id == Employee.id)
            .join(Department, Schedule.department_id == Department.id)
            .join(Shift, Schedule.shift_id == Shift.id)
            .filter(
                and_(
                    Schedule.roster_type == roster_type,
                    Schedule.schedule_date >= start_date,
                    Schedule.schedule_date <= end_date,
                    Employee.user_id == self.user_id,
                )
            )
        )
        if batch_id:
            q = q.filter(Schedule.generation_batch == batch_id)

        q = q.order_by(Schedule.schedule_date, Department.name, Employee.name)
        results = []
        for sched, emp, dept, shift in q.all():
            results.append({
                "date": sched.schedule_date.isoformat(),
                "employee_name": emp.name,
                "role": emp.role,
                "department_name": dept.name,
                "shift_name": shift.name,
                "start_time": shift.start_time,
                "end_time": shift.end_time,
                "ai_score": sched.ai_score,
                "reason": sched.reason,
                "is_manual_override": sched.is_manual_override,
            })
        return results
